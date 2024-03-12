from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import time
import bs4
import os
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import nltk
import pdfplumber
nltk.download('averaged_perceptron_tagger')
from nltk.tokenize import word_tokenize, sent_tokenize
from nltk.tag import pos_tag
nltk.download('punkt')
from nltk.tokenize import sent_tokenize
from nltk.tokenize.punkt import PunktTrainer
from nltk.tokenize.punkt import PunktSentenceTokenizer , PunktTrainer

def extraire_informations_html(fichier_html, ligne_depart):
    # Lire le fichier HTML
    with open(fichier_html, 'r' , encoding='utf-8') as fichier:
        lignes = fichier.readlines()[ligne_depart - 1:]  # Lire les lignes à partir de la ligne de départ

    # Joindre les lignes en une seule chaîne de caractères
    contenu = ''.join(lignes)

    # Utiliser BeautifulSoup pour analyser le HTML
    soup = BeautifulSoup(contenu, 'html.parser')

    # Trouver la balise <a> "Accouvage"
    balise_accouvage = soup.find('a', text='Z')

    # Démarrer l'analyse à partir de la balise Accouvage
    balises_suivantes = balise_accouvage.find_all_next('a')

    # Extraire les balises <a> et leurs liens href
    informations = []
    for balise_suivante in balises_suivantes:
        texte = balise_suivante.get_text()
        lien = balise_suivante.get('href', '')
        informations.append((texte, lien))

    return informations

# Exemple d'utilisation
fichier_html = 'page_exportee.html'  # Remplacez par le nom de votre fichier HTML exporté
ligne_depart = 967  # Ligne de départ pour l'analyse
informations = extraire_informations_html(fichier_html, ligne_depart)

# Créer un DataFrame pandas à partir des informations extraites
df = pd.DataFrame(informations, columns=['Texte', 'Lien'])

# Exporter le DataFrame dans un fichier Excel
nom_fichier_excel = 'CCN_ref.xlsx'  # Nom du fichier Excel de sortie
df.to_excel(nom_fichier_excel, index=False)


# Charger le fichier Excel
wb = openpyxl.load_workbook('CCN_ref.xlsx')
ws = wb.active

# Parcourir les lignes en partant de la fin jusqu'au début
for row in reversed(range(2, ws.max_row + 1)):
    texte = ws.cell(row, 1).value
    lien = ws.cell(row, 2).value

    # Supprimer les lignes où la première colonne est vide
    if texte is None or texte.strip() == '':
        ws.delete_rows(row)

    # Supprimer les lignes après la valeur "Zoos" dans la première colonne
    if texte == 'Zoos':
        ws.delete_rows(row + 1, ws.max_row - row)

# Remplacer les caractères spécifiques dans la colonne "Texte"
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    cell = row[0]
    if cell.value:
        cell.value = cell.value.replace(':', '')
        cell.value = cell.value.replace('<', 'inf')
        cell.value = cell.value.replace('>', 'sup')
        cell.value = cell.value.replace('≤', 'inf-égal')
        cell.value = cell.value.replace('/', '-')

# Enregistrer les modifications dans le même fichier Excel
wb.save('CCN_ref.xlsx')



#--------------------------------EXPORT HTML------------------------------------------



# Lire le fichier Excel avec les liens href
fichier_excel = 'informations_Bureaux.xlsx'  # Nom du fichier Excel
df = pd.read_excel(fichier_excel)

# Créer le dossier "export_html" s'il n'existe pas
dossier_export = 'export_html'
if not os.path.exists(dossier_export):
    os.makedirs(dossier_export)


# Initialiser le navigateur Selenium
driver = webdriver.Chrome(executable_path='chromedriver.exe')  # Remplacez par le pilote approprié pour votre navigateur (Chrome, Firefox, etc.)

# Se connecter au site avec les identifiants
url_login = "https://securega.gestion-des-acces.fr/login?service=https%3a%2f%2fwww.elnet.fr%2fGa%2fLogin%3ffromHulk%3d%252fdocumentation%252fDocument%253fid%253dY5LSTCC-1"  # Remplacez par l'URL de connexion
identifiant = 'ELG7KLLSY'
mot_de_passe = '8KK9SD'


# Naviguer vers la page de connexion
driver.get(url_login)
time.sleep(5)
# Remplir les champs d'identification
driver.find_element('id','username').send_keys(identifiant)  # Remplacez 'identifiant_field_id' par l'ID du champ d'identification réel
driver.find_element('id','soumettre').click()
driver.find_element('id','password').send_keys(mot_de_passe)  # Remplacez 'mot_de_passe_field_id' par l'ID du champ de mot de passe réel

# Soumettre le formulaire de connexion
driver.find_element('id','soumettre').click() # Remplacez 'connexion_button_id' par l'ID du bouton de connexion réel

# Accepter les cookies s'il y a une fenêtre contextuelle
time.sleep(4)  # Attendre 2 secondes pour que la fenêtre contextuelle s'affiche (ajustez le délai si nécessaire)
try:
    driver.find_element('id','_tealiumModalClose').click()  # Remplacez 'accepter_button_id' par l'ID du bouton d'acceptation des cookies réel
    time.sleep(3)
except:
    pass

try:
    driver.find_element('id','didomi-notice-agree-button').click()  # Remplacez 'accepter_button_id' par l'ID du bouton d'acceptation des cookies réel
    #time.sleep(3)
except:
    pass

# Parcourir chaque ligne du DataFrame
for _, row in df.iterrows():
    texte = row['Texte']
    lien = row['Lien']

    # Naviguer vers le lien href
    driver.get(lien)

    # Faire défiler la page pour que toutes les données apparaissent
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)  # Attendre un court instant pour permettre le chargement des données supplémentaires
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    # Récupérer le contenu HTML généré par JavaScript
    contenu = driver.execute_script("return document.getElementsByTagName('html')[0].innerHTML")

    # Enregistrer le contenu dans un fichier HTML
    nom_fichier_html = f"{texte}.html"  # Nom du fichier HTML basé sur le texte
    chemin_fichier_html = os.path.join(dossier_export, nom_fichier_html)
    with open(chemin_fichier_html, 'w', encoding='utf-8') as fichier:
        fichier.write(contenu)

    print(f"Le fichier {nom_fichier_html} a été enregistré dans le dossier {dossier_export}.")

# Fermer le navigateur Selenium
driver.quit()



#-------------------------------Code NAF---------------------------------------------------------------


# Chemin vers le fichier Excel de sortie
fichier_excel = os.path.join(os.getcwd(), 'CCN_Bureaux.xlsx') 

fichier_html_url = os.path.join(os.getcwd(), "export_html\Bureaux d'études techniques.html") 
# Ouvrir le fichier HTML
with open(fichier_html_url, encoding='utf-8') as file:
    soup = BeautifulSoup(file, 'html.parser')

# Fonction pour vérifier si une phrase contient la mention "NAF"
def NAF(sentence):
   words = word_tokenize(sentence.lower())
   tagged_words = pos_tag(words)

   for word, tag in tagged_words:
        if 'naf' in word:
           return True
   return False


 

# Extraire les tableaux contenant la colonne "base conventionnelle"
tables_with_NAF = []


for table in soup.find_all('table'):
   headers = table.find_all('th')
   headers_text = [header.text.strip() for header in headers]


   for header_text in headers_text:
       sentences = sent_tokenize(header_text)
       for sentence in sentences:
           if NAF(sentence):
               tables_with_NAF.append(table)
               break

# Extraire les données des tableaux identifiés
data = []

for table in tables_with_NAF:
   table_data = []
   for row in table.find_all('tr'):
       row_data = []
       for cell in row.find_all('td'):
           row_data.append(cell.text.strip())
       table_data.append(row_data)

# Extraire les données du tableau
data = []
rows = table.find_all('tr')

# Parcourir les lignes du tableau
for row in rows:
    cells = row.find_all('td')
    if len(cells) >= 2:
        code_naf = cells[0].text.strip()
        activite = cells[1].text.strip()
        # Ignorer les lignes qui contiennent des données non pertinentes
        #if code_naf and activite and not code_naf.startswith('Convention collective') and not code_naf.startswith('Date d\'arrêt des te'):
        data.append([code_naf, activite])


# Créer un dataframe pandas avec les données
df = pd.DataFrame(data, columns=['Code NAF', 'Activité'])
# Sauvegarder le DataFrame dans un fichier Excel
df.to_excel(fichier_excel, index=False, sheet_name='Code NAF', engine='xlsxwriter')




#-------------------------------Maintien du salaire------------------------------------------------------------




# Chemin vers le fichier PDF
fichier_pdf = os.path.join(os.getcwd(), "docpdf.pdf")

# Rechercher la colonne "Maintien du salaire" dans le PDF
target_column = "Maintien du salaire - (IJSS + régime de prévoyance)"

# Ouvrir le fichier PDF avec pdfplumber
with pdfplumber.open(fichier_pdf) as pdf:
    # Parcourir les pages du PDF
    for page in pdf.pages:
        # Extraire le texte de la page
        page_text = page.extract_text()
        
        # Rechercher la position de la colonne "Maintien du salaire"
        if target_column in page_text:
            # Extraire le tableau à partir de la position trouvée
            table = page.extract_table()
            
            # Créer un DataFrame à partir du tableau
            df = pd.DataFrame(table[1:], columns=table[0])
            
            # Sauvegarder le DataFrame dans un fichier Excel
            with pd.ExcelWriter(fichier_excel, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name='Maintien du salaire', index=False)
        
            print("Tableau extrait et sauvegardé avec succès dans le fichier Excel.")
            break
    else:
        print("Aucun tableau contenant la colonne 'Maintien du salaire' n'a été trouvé dans le fichier PDF.")



#-------------------------------Bénéficiaire-------------------------------------------------------------




# Ouvrir le fichier HTML
with open(fichier_html_url, encoding='utf-8') as file:
    soup = BeautifulSoup(file, 'html.parser')

text = soup.get_text()

#tokenize.sent_tokenize(text)

trainer = PunktTrainer()

trainer.train(text, finalize=False, verbose=True)

tokenizer = PunktSentenceTokenizer(trainer.get_params())

sentences = tokenizer.tokenize(text)

for paragraphe in sentences:
    print(paragraphe)


# Find the position of the title "Régime de prévoyance"
start_index_regime = text.find("Régimes de prévoyance") if "Régimes de prévoyance" in text else text.find("Régime de prévoyance")

# Find the position of the title "bénéficiaire" under the "Régime de prévoyance" section
start_index_beneficiaire = text.find("Bénéficiaires", start_index_regime)

org = text.find("Organisme", start_index_beneficiaire)
cot = text.find("Cotisations", start_index_beneficiaire)
if cot < org:
    end_index_benif = cot
else:
    end_index_benif = org



if start_index_beneficiaire != -1 and end_index_benif != -1:
    # Extract the text between the "bénéficiaire" and "Cotisations" points
    text_between_titles = text[start_index_beneficiaire: end_index_benif]
    
    # Tokenize the sentences in the extracted text
    sentences = sent_tokenize(text_between_titles, language='french')
    formatted_text = '\n'.join(sentences)

    # Create a DataFrame to hold the sentences
    df_beneficiaire = pd.DataFrame([formatted_text], columns=['Bénéficiaires'])

    # Sauvegarder le DataFrame dans un fichier Excel
    with pd.ExcelWriter(fichier_excel, engine='openpyxl', mode='a') as writer:
        df_beneficiaire.to_excel(writer, sheet_name='Bénéficiaires', index=False)

else:
    print("n'existe pas")




#----------------------------------------Cotisation----------------------------------------------------------




# Find the position of the title "Régime de prévoyance"
start_index_regime = text.find("Régimes de prévoyance") if "Régimes de prévoyance" in text else text.find("Régime de prévoyance")

# Find the position of the title "bénéficiaire" under the "Régime de prévoyance" section
start_index_Cotisations = text.find("Cotisations", start_index_regime)

# Find the position of the word "Cotisations" after the "bénéficiaire" title
end_index_Cotisations = text.find("Prestations", start_index_Cotisations) 


if start_index_beneficiaire != -1 and end_index_Cotisations != -1:
    # Extract the text between the "bénéficiaire" and "Cotisations" points
    text_between_titles = text[start_index_Cotisations :end_index_Cotisations]
    # Tokenize the sentences in the extracted text
    sentences = sent_tokenize(text_between_titles, language='french')

    
    # Concatenate the sentences with appropriate formatting
    formatted_text_cotisation = '\n'.join(sentences)
    
    # Create a DataFrame to hold the sentences
    df_cotisation = pd.DataFrame([formatted_text_cotisation], columns=['Cotisations'])

    # Sauvegarder le DataFrame dans un fichier Excel
    with pd.ExcelWriter(fichier_excel, engine='openpyxl', mode='a') as writer:
        df_cotisation.to_excel(writer, sheet_name='Cotisation', index=False)
else:
    print("n'existe pas")




#---------------------------------Préstation (Minima de garanties)----------------------------------------





# Find the position of the title "Régime de prévoyance"
start_index_regime = text.find("Régimes de prévoyance") if "Régimes de prévoyance" in text else text.find("Régime de prévoyance")

# Find the position of the title "bénéficiaire" under the "Régime de prévoyance" section
start_index_beneficiaire = text.find("Prestations", start_index_regime)

# Find the position of the word "Cotisations" after the "bénéficiaire" title
end_index = text.find("Incidences", start_index_regime) 

if start_index_beneficiaire != -1 and end_index != -1:
    # Extract the text between the "bénéficiaire" and "Cotisations" points
    text_between_titles = text[start_index_beneficiaire + len(" Prestations "):end_index]
    
    # Tokenize the sentences in the extracted text
    sentences = sent_tokenize(text_between_titles, language='french')
    
    # Train the Punkt tokenizer
    trainer = PunktTrainer()
    trainer.train(text_between_titles, finalize=False, verbose=True)
    tokenizer = trainer.get_params()
    
    # Concatenate the sentences with appropriate formatting
    formatted_text_prestation = '\n'.join(sentences)
    
    # Create a DataFrame to hold the sentences
    df_prestation = pd.DataFrame([formatted_text_prestation], columns=['Minima de garanties'])

    # Sauvegarder le DataFrame dans un fichier Excel
    with pd.ExcelWriter(fichier_excel, engine='openpyxl', mode='a') as writer:
        df_prestation.to_excel(writer, sheet_name='Minima de garanties', index=False)
else:
    print("ereeur.")
