from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook
import pandas as pd
import time
import os
import logging

def set_up_logging(path: str):
    # Configuration des logs
    logging.basicConfig(
    level=logging.DEBUG,  # Niveau minimum des logs à afficher
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        #logging.StreamHandler(),  # Affiche les logs dans la console
        logging.FileHandler(os.path.join(path, 'log_donwload_html_content.txt')) 
            ])

def extraire_informations_html(fichier_html, ligne_depart):
    # Lire le fichier HTML
    with open(fichier_html, 'r' , encoding='utf-8') as fichier:
        # Lire les lignes à partir de la ligne de départ
        lignes = fichier.readlines()[ligne_depart - 1:] 

    # Joindre les lignes en une seule chaîne de caractères
    contenu = ''.join(lignes)

    # Utiliser BeautifulSoup pour analyser le HTML
    soup = BeautifulSoup(contenu, 'html.parser')

    # Trouver la balise <a> "Accouvage"
    balise_accouvage = soup.find('a', string='Z')

    # Démarrer l'analyse à partir de la balise Accouvage
    balises_suivantes = balise_accouvage.find_all_next('a')

    # Extraire les balises <a> et leurs liens href
    informations = []
    for balise_suivante in balises_suivantes:
        texte = balise_suivante.get_text()
        lien = balise_suivante.get('href', '')
        informations.append((texte, lien))

    return informations

def donwload_workbook(
        path_excel_path: str
        ):
    workbook = load_workbook(path_excel_path)
    return workbook

def delete_row_from_workbook(
        worksheet,
        word_to_delete_after: str = 'Zoos'
        ):
    # Parcourir les lignes en partant de la fin jusqu'au début
    for row in reversed(range(2, worksheet.max_row + 1)):
        text = worksheet.cell(row, 1).value

        # Supprimer les lignes où la première colonne est vide
        if text is None or text.strip() == '':
            worksheet.delete_rows(row)

        # Supprimer les lignes après la valeur "Zoos" dans la première colonne
        if text == word_to_delete_after:
            worksheet.delete_rows(row + 1, worksheet.max_row - row)
    return worksheet

def clean_row_of_workbook_for_colonne_texte(worksheet):
    for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=1):
        cell = row[0]
        if cell.value:
            cell.value = cell.value.replace(':', '')
            cell.value = cell.value.replace('<', 'inf')
            cell.value = cell.value.replace('>', 'sup')
            cell.value = cell.value.replace('≤', 'inf-égal')
            cell.value = cell.value.replace('/', '-')
    return worksheet

def create_dataframe(path_excel: str) -> pd.DataFrame:
    dataframe_created = pd.read_excel(path_excel)
    return dataframe_created

def get_pause_for_seconds(time_to_pause: int):
    time.sleep(time_to_pause)

def get_initialization(
        path_executable: str
        ) -> webdriver.Chrome:
    """
    Create the driver object
    Remplacez par le pilote approprié pour votre navigateur (Chrome, Firefox, etc.)
    """
    driver = webdriver.Chrome(executable_path=path_executable) 
    return driver 

def get_connection_to_website(
        driver: webdriver.Chrome,
        url_login: str,
        identifiant: str,
        mot_de_passe: str
        ) -> webdriver.Chrome:
    
    # Naviguer vers la page de connexion
    driver.get(url_login)
    get_pause_for_seconds(5)
    # Remplir les champs d'identification
    # Remplacez 'identifiant_field_id' par l'ID du champ d'identification réel
    driver.find_element('id','username').send_keys(identifiant)  
    driver.find_element('id','soumettre').click()
    driver.find_element('id','password').send_keys(mot_de_passe)
    return driver 

def accept_button(
        driver: webdriver.Chrome,
        balise_to_locate_element: str,
        text_to_search: str,
        is_with_pause: bool
        ) -> None:
    """
    driver : webdriver.Chrome to make the action. It have to be Initialized
    Affect balise_to_locate_element by the id of real cookie acceptance buton
    text_to_search: text to find
    is_with_pause: bool that indicate if the user want to have some sleep time of 3s
    """
    try:
        driver.find_element(balise_to_locate_element, text_to_search).click() 
        if is_with_pause:
            get_pause_for_seconds(3)
    except:
        pass

def scroll_all_page(
        driver: webdriver.Chrome
        ):
    # Faire défiler la page pour que toutes les données apparaissent
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        # Attendre un court instant pour permettre le chargement des données supplémentaires
        get_pause_for_seconds(3)  
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

def extract_content(
        driver: webdriver.Chrome
        ) -> str:
    javascript_to_get_content = "return document.getElementsByTagName('html')[0].innerHTML"
    content = driver.execute_script(javascript_to_get_content)
    return content

def save_html_content_in_folder(
        text: str,
        content: str,
        path_export_folder: str
        ) -> None:
    try:
        # Enregistrer le contenu dans un fichier HTML
        nom_fichier_html = f"{text}.html"  # Nom du fichier HTML basé sur le texte
        chemin_fichier_html = os.path.join(path_export_folder, nom_fichier_html)
        with open(chemin_fichier_html, 'w', encoding='utf-8') as fichier:
            fichier.write(content)
        
        print(f"Le fichier {nom_fichier_html} a été enregistré dans le dossier.")
    except ValueError as e:
        raise e 


